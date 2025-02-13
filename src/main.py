# Copyright 2023 agenius666
# GitHub: https://github.com/agenius666/Adjustment-Entries-Automator
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

import os
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
from datetime import datetime
import warnings
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import threading
import queue
import subprocess

def edit_dataframe(df, group_column):
    """根据指定列分组，并在每一组之间插入空行，返回处理后的 DataFrame"""
    edited_adj = pd.DataFrame(columns=df.columns)  
    grouped = df.groupby(group_column)  # 根据指定列分组

    for group_name, group_data in grouped:
        # 将当前组的数据添加到结果 DataFrame 中
        edited_adj = pd.concat([edited_adj, group_data], ignore_index=True)
        # 在当前组后面插入空行
        edited_adj = pd.concat([edited_adj, pd.DataFrame([{col: '' for col in df.columns}])], ignore_index=True)

    # 始终不录入第一列的内容
    edited_adj = edited_adj.iloc[:, 1:]
    return edited_adj  # 返回处理后的 DataFrame

def update_workbook(file_path):
    """刷新工作簿以触发公式重新计算"""
    wb = load_workbook(file_path)
    wb.save(file_path)

class AdjustmentProcessor:
    def __init__(self, folder, log_queue, stop_event, tb_path_file, adj_file):
        self.folder = folder
        self.log_queue = log_queue
        self.stop_event = stop_event
        self.success_count = 0
        self.error_count = 0
        self.results = []
        self.total_tasks = 0
        self.tb_path_file = tb_path_file
        self.adj_file = adj_file

    def process(self):
        try:
            tb_path = pd.read_excel(self.tb_path_file, header=2)
            adj = pd.read_excel(self.adj_file)
            adj_comp = sorted(map(str, adj['账套名称'].unique()))
            self.total_tasks = len(adj_comp) + 3  # 账套处理 + 3个合并类型

            log_file = os.path.join(self.folder, "日志.txt")
            with open(log_file, 'w', encoding='utf-8') as f:  # 每次操作前重新生成日志文件
                f.write("处理日志:\n")

            self.results = [('账套名称', '调整前未分配利润', '调整前资产负债表平否', '调整后未分配利润', '调整后资产负债表平否')]

            for comp in adj_comp:
                if self.stop_event.is_set():
                    self.log('处理已终止', 'warning')
                    break

                try:
                    self.log(f'正在处理 {comp}...')
                    tb_file = tb_path.loc[tb_path['账套名称'] == comp, '相对路径'].iloc[0]
                    tb_path_full = os.path.join(self.folder, tb_file)
                    adj_data = adj.loc[adj['账套名称'] == comp]
                    adj_processed = edit_dataframe(adj_data, '编号')

                    wb = load_workbook(tb_path_full, data_only=True)
                    re1 = wb['资产负债表']['H45'].value
                    check1 = wb['资产负债表']['A3'].value
                    wb.close()

                    wb = load_workbook(tb_path_full)
                    ws = wb['调整分录']
                    last_row = ws.max_row
                    num_rows = len(adj_processed.index)
                    rows_to_insert = max(0, num_rows - last_row + 4)

                    num_cols = len(adj_processed.columns)  # 获取调整分录的列数
                    for row in ws.iter_rows(min_row=6, max_row=last_row, min_col=1, max_col=num_cols):
                        for cell in row:
                            cell.value = None

                    if rows_to_insert > 0:
                        ws.insert_rows(7, amount=rows_to_insert)
                    for i, row in enumerate(adj_processed.values, start=7):
                        for j, value in enumerate(row, start=1):
                            ws.cell(row=i, column=j, value=value)
                    wb.save(tb_path_full)
                    wb.close()

                    wb = load_workbook(tb_path_full, data_only=True)
                    re2 = wb['资产负债表']['H45'].value
                    check2 = wb['资产负债表']['A3'].value
                    wb.close()

                    self.results.append((comp, re1, check1, re2, check2))
                    self.success_count += 1
                    self.log(f'{comp} 处理完成', 'success')

                except Exception as e:
                    self.error_count += 1
                    self.log(f'{comp} 处理失败: {str(e)}', 'error')
                    with open(log_file, 'a', encoding='utf-8') as f:
                        f.write(f'{comp} 调整分录录入有误，Check: {e}\n')

            if not self.stop_event.is_set():
                result_file = os.path.join(self.folder, f'录入结果_{datetime.now().strftime("%Y-%m-%d_%H-%M-%S")}.xlsx')
                new_wb = Workbook()
                new_ws = new_wb.active
                new_ws.title = "汇总数据"
                for row in self.results:
                    new_ws.append(row)
                new_wb.save(result_file)
                new_wb.close()

                self.log('开始刷新合并TB...')
                for merge_type in ['小合并', '中合并', '大合并']:
                    if self.stop_event.is_set():
                        self.log('处理已终止', 'warning')
                        break

                    for tb_file in tb_path[tb_path['类型'] == merge_type]['相对路径']:
                        try:
                            update_workbook(os.path.join(self.folder, tb_file))
                            self.log(f'{merge_type} {tb_file} 刷新完成', 'success')
                        except Exception as e:
                            self.log(f'{merge_type} {tb_file} 刷新失败: {str(e)}', 'error')

                self.log('所有操作已完成！', 'success')
                self.log_queue.put(('COMPLETE', self.success_count, self.error_count, result_file))

        except Exception as e:
            self.log(f'全局错误: {str(e)}', 'error')
            self.log_queue.put(('ERROR', str(e)))

    def log(self, message, tag='info'):
        self.log_queue.put(('LOG', message, tag))

class AdjustmentApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Adjusting Entries Automator - 1.0.0")
        self.log_queue = queue.Queue()
        self.stop_event = threading.Event()
        self.result_file = None
        self.tb_path_file = None
        self.adj_file = None
        self.setup_ui()
        self.check_log_queue()

    def setup_ui(self):
        # 样式配置
        style = ttk.Style()
        style.configure('TButton', padding=6, font=('微软雅黑', 10))
        style.configure('TLabel', font=('微软雅黑', 9))
        style.configure('Header.TLabel', font=('微软雅黑', 10, 'bold'))

        # 主布局
        main_frame = ttk.Frame(self.root, padding=10)
        main_frame.pack(fill=tk.BOTH, expand=True)

        # 路径选择
        path_frame = ttk.Frame(main_frame)
        path_frame.pack(fill=tk.X, pady=5)
        ttk.Label(path_frame, text="工作文件夹:", style='Header.TLabel').pack(side=tk.LEFT)
        self.path_entry = ttk.Entry(path_frame, width=50)
        self.path_entry.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
        ttk.Button(path_frame, text="浏览...", command=self.browse_folder).pack(side=tk.LEFT)
        ttk.Button(path_frame, text="生成配置文件", command=self.generate_config_files).pack(side=tk.LEFT, padx=5)

        # 文件选择
        file_frame = ttk.Frame(main_frame)
        file_frame.pack(fill=tk.X, pady=5)
        ttk.Label(file_frame, text="路径表文件:", style='Header.TLabel').pack(side=tk.LEFT)
        self.tb_path_entry = ttk.Entry(file_frame, width=50)
        self.tb_path_entry.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
        ttk.Button(file_frame, text="浏览...", command=self.browse_tb_path_file).pack(side=tk.LEFT)

        ttk.Label(file_frame, text="调整分录文件:", style='Header.TLabel').pack(side=tk.LEFT, padx=10)
        self.adj_entry = ttk.Entry(file_frame, width=50)
        self.adj_entry.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
        ttk.Button(file_frame, text="浏览...", command=self.browse_adj_file).pack(side=tk.LEFT)

        # 日志区域
        log_frame = ttk.LabelFrame(main_frame, text="处理日志", padding=5)
        log_frame.pack(fill=tk.BOTH, expand=True)

        self.log_text = tk.Text(log_frame, height=15, state=tk.DISABLED, wrap=tk.WORD)
        scrollbar = ttk.Scrollbar(log_frame, command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=scrollbar.set)

        self.log_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # 标签配置
        self.log_text.tag_config('info', foreground='black')
        self.log_text.tag_config('success', foreground='green')
        self.log_text.tag_config('error', foreground='red')
        self.log_text.tag_config('warning', foreground='orange')

        # 按钮区域
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(pady=10)
        self.start_btn = ttk.Button(btn_frame, text="开始处理", command=self.start_processing)
        self.start_btn.pack(side=tk.LEFT, padx=5)
        self.stop_btn = ttk.Button(btn_frame, text="终止处理", command=self.stop_processing, state=tk.DISABLED)
        self.stop_btn.pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="退出", command=self.root.quit).pack(side=tk.RIGHT, padx=5)

    def browse_folder(self):
        folder = filedialog.askdirectory()
        if folder:
            self.path_entry.delete(0, tk.END)
            self.path_entry.insert(0, folder)

    def browse_tb_path_file(self):
        file = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if file:
            self.tb_path_entry.delete(0, tk.END)
            self.tb_path_entry.insert(0, file)
            self.tb_path_file = file

    def browse_adj_file(self):
        file = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if file:
            self.adj_entry.delete(0, tk.END)
            self.adj_entry.insert(0, file)
            self.adj_file = file

    def generate_config_files(self):
        """生成配置文件和文件夹"""
        folder = self.path_entry.get()
        if not folder:
            messagebox.showerror("错误", "请先选择工作文件夹！")
            return

        config_folder = os.path.join(folder, '配置文件')
        tb_folder = os.path.join(folder, 'TB文件')

        # 检查配置文件是否已存在
        if os.path.exists(config_folder):
            messagebox.showinfo("提示", "配置文件已存在，无需重新生成！")
            return

        # 创建文件夹
        os.makedirs(config_folder, exist_ok=True)
        os.makedirs(tb_folder, exist_ok=True)

        # 创建调整分录.xlsx
        adj_file = os.path.join(config_folder, '调整分录.xlsx')
        wb = Workbook()
        ws = wb.active
        ws.title = "调整分录"
        ws['A1'] = '账套名称'
        ws['B1'] = '从这里复制你的调整分录表头'

        # 设置单元格的底色
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        ws['A1'].fill = yellow_fill

        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        ws['B1'].fill = red_fill

        wb.save(adj_file)
        wb.close()

        # 创建路径表.xlsx
        path_file = os.path.join(config_folder, '路径表.xlsx')
        wb = Workbook()
        ws = wb.active
        ws.title = "路径表"
        ws['A1'] = '文件夹名称'
        ws['A3'] = '账套名称'
        ws['B3'] = '类型'
        ws['C3'] = '文件名称'
        ws['D3'] = '法人文件夹'
        ws['E3'] = '相对路径'
        ws['E4'] = '=TEXTJOIN("/",TRUE,$B$1,D4,C4)'

        # 设置单元格的底色
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        ws['A3'].fill = yellow_fill

        green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        ws['B1'].fill = green_fill

        wb.save(path_file)
        wb.close()

        messagebox.showinfo("提示", "配置文件和文件夹已生成！")

    def start_processing(self):
        folder = self.path_entry.get()
        if not folder:
            messagebox.showerror("错误", "请先选择工作文件夹！")
            return

        if not self.tb_path_file or not self.adj_file:
            messagebox.showerror("错误", "请先选择路径表和调整分录文件！")
            return

        self.start_btn.config(state=tk.DISABLED)
        self.stop_btn.config(state=tk.NORMAL)
        self.log_text.config(state=tk.NORMAL)
        self.log_text.delete(1.0, tk.END)
        self.log_text.config(state=tk.DISABLED)
        self.stop_event.clear()

        processor = AdjustmentProcessor(folder, self.log_queue, self.stop_event, self.tb_path_file, self.adj_file)
        self.process_thread = threading.Thread(target=processor.process, daemon=True)
        self.process_thread.start()

    def stop_processing(self):
        self.stop_event.set()
        self.stop_btn.config(state=tk.DISABLED)
        self.start_btn.config(state=tk.NORMAL)
        self.log('处理已终止', 'warning')

    def check_log_queue(self):
        try:
            while True:
                msg = self.log_queue.get_nowait()
                if msg[0] == 'LOG':
                    _, message, tag = msg
                    self.append_log(message, tag)
                elif msg[0] == 'COMPLETE':
                    _, success, error, result_file = msg
                    self.show_result(success, error, result_file)
                elif msg[0] == 'ERROR':
                    _, error_msg = msg
                    messagebox.showerror("全局错误", error_msg)
        except queue.Empty:
            pass
        finally:
            self.root.after(100, self.check_log_queue)

    def append_log(self, message, tag='info'):
        self.log_text.config(state=tk.NORMAL)
        self.log_text.insert(tk.END, message + "\n", tag)
        self.log_text.config(state=tk.DISABLED)
        self.log_text.see(tk.END)

    def show_result(self, success, error, result_file):
        self.start_btn.config(state=tk.NORMAL)
        self.stop_btn.config(state=tk.DISABLED)
        self.result_file = result_file

        result_window = tk.Toplevel(self.root)
        result_window.title("处理结果")

        ttk.Label(result_window, text="处理统计", font=('微软雅黑', 12, 'bold')).pack(pady=5)
        ttk.Label(result_window, text=f"成功处理账套数：{success}").pack()
        ttk.Label(result_window, text=f"处理失败账套数：{error}").pack()

        btn_frame = ttk.Frame(result_window)
        btn_frame.pack(pady=10)

        ttk.Button(btn_frame, text="打开结果文件",
                   command=lambda: subprocess.Popen(['start', 'excel', result_file], shell=True)).pack(side=tk.LEFT,
                                                                                                     padx=5)
        ttk.Button(btn_frame, text="查看完整日志",
                   command=lambda: subprocess.Popen(['notepad', os.path.join(self.path_entry.get(), "日志.txt")])).pack(
            side=tk.LEFT, padx=5)

if __name__ == "__main__":
    warnings.simplefilter(action='ignore', category=FutureWarning)
    root = tk.Tk()
    root.geometry("800x600")
    app = AdjustmentApp(root)
    root.mainloop()

